from typing import List, Optional
from fastapi import Depends, FastAPI, Form, HTTPException, Header, Query, Request, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi_jwt_auth import AuthJWT
from fastapi_jwt_auth.exceptions import MissingTokenError, JWTDecodeError
from datetime import datetime
from uuid import uuid4
from pydantic import BaseModel
from pymongo import MongoClient
from openpyxl import Workbook
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
import os
import socket
from schema import RequirementRequest, RequirementResponse, UserCreate, UserLogin
from utils import Settings, build_search_filter, calculate_storage, estimate_bitrate, get_password_hash, recommend_server, update_bitrate, verify_password

hostname = socket.gethostname()
local_ip = socket.gethostbyname(hostname)


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.144.30.1:8000"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

client = MongoClient("mongodb://192.168.1.67:27017")
db = client["redx_vms"]
collection = db["requirements"]
users_collection = db["users"]
EXPORT_DIR = "exports"
TEMPLATE_DIR = "templates"
os.makedirs(EXPORT_DIR, exist_ok=True)
env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))

templates = Jinja2Templates(directory=TEMPLATE_DIR)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(BASE_DIR, ".", "static")
app.mount("/static", StaticFiles(directory=static_dir), name="static")


class User(BaseModel):
    username: str
    email: Optional[str] = None
    full_name: Optional[str] = None
    disabled : Optional[bool] = False
    role : str = 'user'

class UserInDB(User):
    hashed_password : str

class Token(BaseModel):
    access_token : str
    token_type : str

class TokenData(BaseModel):
    username : Optional[str] = None
    role : Optional[str] = None


class AdminUserCreate(BaseModel):
    username: str
    email: str
    full_name: str
    password: str
    role: str = "user"
    disabled: bool = False

class UserListResponse(BaseModel):
    users: List[User]


class UserBase(BaseModel):
    username: str
    email : str
    role : str

class UserCreate(UserBase):
    password : str
    is_active : bool = False
    created_by : Optional[str] = None

class UserUpdate(BaseModel):
    email : Optional[str] = None
    password : Optional[str] = None
    role : Optional[str] = None
    is_active: bool

class UserOut(UserBase):
    is_active: bool


def create_superadmin():
    superadmin = users_collection.find_one({"username": "superadmin"})
    if not superadmin:
        hashed_password = get_password_hash("12345678")
        users_collection.insert_one({
            "username": "superadmin",
            "hashed_password": hashed_password,
            "email": "superadmin@gmail.com",
            "full_name": "Super Admin",
            "disabled": False,
            "role": "superadmin"
        })
create_superadmin()




@AuthJWT.load_config
def get_config():
    return Settings()

def get_current_user(Authorize: AuthJWT = Depends()):
    try:
        Authorize.jwt_required() 
    except MissingTokenError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing access token")
    except JWTDecodeError:
        try:
            Authorize.jwt_refresh_token_required()
            current_user = Authorize.get_jwt_subject()
            new_access_token = Authorize.create_access_token(subject=current_user)
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail='Token Refresh',
                headers={"X-New-Access-Token": new_access_token}
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail='Invalid Token. Please login again.'
            )
    current_user = Authorize.get_jwt_subject()
    return current_user


async def get_admin(Authorize: AuthJWT = Depends()):
    try:
        Authorize.jwt_required()
        username = Authorize.get_jwt_subject()
        user_data = Authorize.get_raw_jwt()
        
        user = users_collection.find_one({"username": username})
        if not user or user.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail='Super Admin privileges required'
            )
            
        return UserInDB(**user)
    
    except MissingTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Please login to access this page"
        )
        
    except JWTDecodeError:
        try:
            Authorize.jwt_refresh_token_required()
            current_user = Authorize.get_jwt_subject()
            new_access_token = Authorize.create_access_token(subject=current_user)
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Token refreshed",
                headers={"X-New-Access-Token": new_access_token}
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid token. Please login again."
            )
        
    
async def get_superadmin(Authorize: AuthJWT = Depends()):
    try:
        Authorize.jwt_required()
        username = Authorize.get_jwt_subject()
        user_data = Authorize.get_raw_jwt()
        roles = ["superadmin", "admin"]
        user = users_collection.find_one({"username": username})
        if not user or user.get("role") not in roles:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail='Super Admin privileges required'
            )
            
        return UserInDB(**user)
    
    except MissingTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Please login to access this page"
        )
        
    except JWTDecodeError:
        try:
            Authorize.jwt_refresh_token_required()
            current_user = Authorize.get_jwt_subject()
            new_access_token = Authorize.create_access_token(subject=current_user)
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Token refreshed",
                headers={"X-New-Access-Token": new_access_token}
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid token. Please login again."
            )

@app.get('/register', response_class=HTMLResponse)
def register(request: Request):
    context = {
        'request': request,
        "local_ip": local_ip
    }
    return templates.TemplateResponse("register.html", context)

@app.post('/register', response_model=User)
async def register(user: UserCreate):
    existing_user = users_collection.find_one({"username": user.username})
    if existing_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='username already registered')
    hashed_password = get_password_hash(user.password)
    user_dict = user.dict()
    user_dict['hashed_password'] = hashed_password
    del user_dict['password']
    users_collection.insert_one(user_dict)
    return user_dict

@app.get("/login")
async def login(request:Request):
    context = {
        "request":request,
        "local_ip": local_ip
    }
    return templates.TemplateResponse("login.html", context)

@app.post('/login')
async def login(username: str = Form(...), password: str = Form(...), Authorize: AuthJWT = Depends()):
    user = users_collection.find_one({"username": username})
    if not user or not verify_password(password, user['hashed_password']):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Incorrect username or password')

    access_token = Authorize.create_access_token(subject=username)
    refresh_token = Authorize.create_refresh_token(subject=username)

    users_collection.update_one({"username": username}, {"$set": {"refresh_token": refresh_token}})
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "role": user.get('role'),
        "token_type": "bearer",
        "redirect_url": "/" if user['role'] == 'superadmin' else '/'
    }


@app.get('/logout')
def logout(Authorize: AuthJWT = Depends()):
    try:
        current_user = Authorize.get_jwt_subject()
        users_collection.find_one({"username": current_user}, {"$unset": {"refresh_token": ""}})
        Authorize.unset_jwt_cookies()
        return {"msg": "Successfully logged out"}
    except Exception:
        return {"msg": "Successfully logged out"}

@app.get('/protected')
async def protected_route(current_user: User = Depends(get_current_user)):
    return {"message": f"Hello {current_user}"}

@app.get('/admin-only')
async def admin_route(superadmin: User = Depends(get_superadmin)):
    return {"Message": "Welcome Super Admin!"}

@app.post("/refresh")
def refresh(Authorize: AuthJWT = Depends(), authorization: str = Header(None)):
    try:
        Authorize.jwt_refresh_token_required()
        current_user = Authorize.get_jwt_subject()

        user = users_collection.find_one({"username": current_user})
        if not user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="User not found"
            )
        if authorization is None or not authorization.startswith("Bearer "):
            raise HTTPException(status_code=401, detail="No refresh token provided")
        
        authorization_refresh_token = authorization.split(" ")[1]
        
        if user.get("refresh_token") != authorization_refresh_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail='Invalid refresh token'
            )

        new_access_token = Authorize.create_access_token(subject=current_user)
        
        return {
            "access_token": new_access_token,
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, 
            detail="Invalid refresh token"
        )

@app.get("/", response_class=HTMLResponse)
def home(request: Request):

    context = {
        "request": request,
        "local_ip": local_ip,
        "is_authenticated":""
    }
    return templates.TemplateResponse("index1.html", context)


@app.post("/requirement", response_model=RequirementResponse)
def create_requirement(req: RequirementRequest, current_user: User = Depends(get_current_user)):
    try:
        update_bitrate(req)

        uid = str(uuid4())
        bandwidth = 0
        for cam in req.camera_configs:
            if cam.bitrate_kbps:
                bitrate_mbps = cam.bitrate_kbps / 1000
            else:
                bitrate_mbps = estimate_bitrate(cam.resolution, cam.fps, cam.codec)         
            total_bitrate = bitrate_mbps * cam.qty
            bandwidth+= total_bitrate

        max_retention = max(c.retention_days for c in req.camera_configs)
        avg_record_hour = max(c.record_hour for c in req.camera_configs)
        camera_configs = [cam.dict() for cam in req.camera_configs]
        doc = {
            "_id": uid,
            "customer_name": req.customer_name,
            "project_name": req.project_name,
            "location": req.location,
            "assigned_person": req.assigned_person,
            "camera_configs": camera_configs,
            "created_at": datetime.utcnow(),
            "bandwidth": bandwidth,
            "storage_tb": calculate_storage(total_bitrate, max_retention, avg_record_hour), 
            "server_spec": recommend_server(sum(cam.qty for cam in req.camera_configs), total_bitrate, round(total_bitrate, 2), max_retention, avg_record_hour, camera_configs),
            "created_by": current_user
        }
        collection.insert_one(doc)
        return RequirementResponse(**doc, id=uid)

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/requirement/{id}", response_model=RequirementResponse)
def get_requirement(id: str, Authorize: AuthJWT=Depends()):
    Authorize.jwt_required()
    doc = collection.find_one({"_id": id})
    if not doc:
        raise HTTPException(status_code=404, detail="Requirement not found")
    return RequirementResponse(**doc, id=doc.pop("_id"))


@app.get("/requirement/{id}/export/xlsx")
def export_excel(id: str):
    doc = collection.find_one({"_id": id})
    if not doc:
        raise HTTPException(status_code=404, detail="Requirement not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "REDX VMS Requirement"

    ws.append(["Customer", doc["customer_name"]])
    ws.append(["Project", doc["project_name"]])
    ws.append(["Location", doc["location"]])
    ws.append(["Assigned Person", doc["assigned_person"]])
    ws.append([])
    ws.append(["Camera Brand", "Resolution", "FPS", "Codec", "Record Hour", "Retention Days", "Quantity"])
    for cam in doc["camera_configs"]:
        ws.append([
            cam["name"], cam["resolution"], cam["fps"], cam["codec"],
            cam["record_hour"], cam["retention_days"], cam["qty"]
        ])
    ws.append([])
    ws.append(["Estimated Bitrate (Mbps)", doc["bandwidth"]])
    ws.append(["Estimated Storage (TB)", doc["storage_tb"]])
    ws.append(["Server CPU", doc["server_spec"]["cpu"]])
    ws.append(["RAM (GB)", doc["server_spec"]["ram_gb"]])
    ws.append(["HDD Total (TB)", doc["server_spec"]["hdd_tb"]])
    ws.append(["NIC Ports", doc["server_spec"]["nic"]])

    path = os.path.join(EXPORT_DIR, f"requirement_{id}.xlsx")
    wb.save(path)
    return FileResponse(path, filename=f"redx_report_{id}.xlsx")


@app.get("/requirement/{id}/export/pdf")
async def export_pdf(
    id: str,
    token: str = Query(..., description="JWT token for authentication"),
    Authorize: AuthJWT = Depends()
):
    try:
        Authorize.jwt_required("access", token=token)
        
        current_user = Authorize.get_jwt_subject()
        jwt_claims = Authorize.get_raw_jwt()
        doc = collection.find_one({"_id": id})
        if not doc:
            raise HTTPException(status_code=404, detail="Requirement not found")

        for cam in doc.get("camera_configs", []):
            quantity = cam.get('qty', 1)
            bitrate_kbps = cam.get('bitrate_kbps', 0)
            cam["bandwidth"] = round(bitrate_kbps * quantity / 1024, 2)

        template = env.get_template("report_template.html")
        html_out = template.render(
            data=doc,
            created=datetime.now().strftime("%d-%b-%Y"),
            user=current_user
        )
        os.makedirs(EXPORT_DIR, exist_ok=True)
        pdf_path = os.path.join(EXPORT_DIR, f"requirement_{id}.pdf")
        
        HTML(string=html_out).write_pdf(pdf_path)
        
        return FileResponse(
            pdf_path,
            filename=f"redx_report_{id}.pdf",
            media_type='application/pdf'
        )
    except Exception as e:
        error_detail = "Invalid or expired token" if "token" in str(e).lower() else str(e)
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"PDF export failed: {error_detail}"
        )

@app.get("/requirement/list/")
def list_all_requirements(current_user: User = Depends(get_current_user)):

    user = users_collection.find_one({"username": current_user})
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='User not found')
    results = []
    role = user["role"]

    if role == 'user':
        query = {"created_by": current_user}
    elif role == 'admin':
        users_created_by_admin = users_collection.find({"created_by": current_user})
        users_created_by_admin = list(users_created_by_admin)
        created_usernames = [user['username'] for user in users_created_by_admin]
        created_usernames.append(current_user)
        query = {"created_by": {"$in": created_usernames}}
    elif role == "superadmin":
        query = {}
    for doc in collection.find(query).sort("created_at", -1):
        total_qty = sum(cam.get("qty", 1) for cam in doc.get('camera_configs', []))
        total_bitrate = doc.get('bandwidth', 0)
        total_bandwidth = round(total_bitrate * total_qty / 1024, 2)

        results.append({
            "id": str(doc["_id"]),
            "project_name": doc["project_name"],
            "customer_name": doc["customer_name"],
            "location": doc.get("location", ""),
            "assigned_person": doc.get("assigned_person", ""),
            "created_at": doc.get("created_at", "")
        })
    return results


@app.get("/requirement/export/all/xlsx")
def export_all_excel(Authorize: AuthJWT = Depends()):
    try:
        query = {}
        Authorize.jwt_required()
        current_user = Authorize.get_jwt_subject()
        user = users_collection.find_one({"username": current_user})
        if user.get('role') == 'user':
            docs = list(collection.find({'created_by': current_user}))
            if not docs:
                raise HTTPException(status_code=404, detail="No requirements found")
        elif user.get('role') == 'admin':
            users_created_by_admin = users_collection.find({"created_by": current_user})
            users_created_by_admin = list(users_created_by_admin)
            created_usernames = [user['username'] for user in users_created_by_admin]
            created_usernames.append(current_user)
            query = {"created_by": {"$in": created_usernames}}
            docs = list(collection.find(query))
            if not docs:
                raise HTTPException(status_code=404, detail="No requirements found")
        else:
            docs = list(collection.find())
            if not docs:
                raise HTTPException(status_code=404, detail="No requirements found")

        wb = Workbook()
        ws = wb.active
        ws.title = "All REDX Requirements"

        ws.append([
            "Customer", "Project", "Location", "Assigned", "Bitrate (Mbps)",
            "Storage (TB)", "Camera Brand", "Resolution", "FPS", "Codec",
            "Record Hour", "Retention Days", "Qty"
        ])

        for doc in docs:
            for cam in doc["camera_configs"]:
                ws.append([
                    doc["customer_name"],
                    doc["project_name"],
                    doc["location"],
                    doc["assigned_person"],
                    doc["bandwidth"],
                    doc["storage_tb"],
                    cam["name"],
                    cam["resolution"],
                    cam["fps"],
                    cam["codec"],
                    cam["record_hour"],
                    cam["retention_days"],
                    cam["qty"]
                ])

        filename = f"redx_all_requirements_{datetime.now().strftime('%Y%m%d')}.xlsx"
        path = os.path.join(EXPORT_DIR, filename)
        wb.save(path)
        
        return FileResponse(
            path,
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@app.get("/requirement/export/filtered/xlsx")
def export_filtered_excel(
    request: Request,
    query: str = Query(None),
    customer_name: str = Query(None),
    project_name: str = Query(None),
    location: str = Query(None),
    assigned_person: str = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    Authorize: AuthJWT = Depends()
):
    try:
        try:
            Authorize.jwt_required()
        except:
            token = request.query_params.get("token")
            if token:
                Authorize.jwt_required("access", token=token)
            else:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED, detail='Missing Token'
                )
        start_date_dt = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
        end_date_dt = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
            
        search_filter = build_search_filter(
            query=query,
            customer_name=customer_name,
            project_name=project_name,
            location=location,
            assigned_person=assigned_person,
            start_date=start_date_dt,
            end_date=end_date_dt
        )

        current_user = Authorize.get_jwt_subject()
        user = users_collection.find_one({"username": current_user})
        if user.get('role') == 'user':
            docs = list(collection.find(search_filter | {'created_by': current_user}).sort("created_at", -1))
            if not docs:
                raise HTTPException(status_code=404, detail="No requirements found")
            
        elif user.get('role') == 'admin':
            users_created_by_admin = users_collection.find({"created_by": current_user})
            users_created_by_admin = list(users_created_by_admin)
            created_usernames = [user['username'] for user in users_created_by_admin]
            created_usernames.append(current_user)
            created_filter = {"created_by": {"$in": created_usernames}}
            docs = list(collection.find(search_filter | created_filter).sort("created_at", -1))
            if not docs:
                raise HTTPException(status_code=404, detail="No requirements found")
        else:
            docs = list(collection.find(search_filter).sort("created_at", -1))
            if not docs:
                raise HTTPException(status_code=404, detail="No requirements found matching the filters")

        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered REDX Requirements"

        ws.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        if query:
            ws.append(["Search Query", query])
        if customer_name:
            ws.append(["Customer Filter", customer_name])
        if project_name:
            ws.append(["Project Filter", project_name])
        if location:
            ws.append(["Location Filter", location])
        if assigned_person:
            ws.append(["Assigned Person Filter", assigned_person])
        if start_date or end_date:
            date_range = []
            if start_date:
                date_range.append(f"From: {start_date}")
            if end_date:
                date_range.append(f"To: {end_date}")
            ws.append(["Date Range", " ".join(date_range)])
        ws.append([])

        ws.append([
            "Customer", "Project", "Location", "Assigned", "Created At",
            "Camera Brand", "Resolution", "FPS", "Codec", "Record Hour", 
            "Retention Days", "Qty", "Bitrate (Mbps)", "Storage (TB)"
        ])

        for doc in docs:
            for cam in doc.get("camera_configs", []):
                created_at = doc.get("created_at", "")
                if isinstance(created_at, datetime):
                    created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
                
                ws.append([
                    doc["customer_name"],
                    doc["project_name"],
                    doc.get("location", ""),
                    doc.get("assigned_person", ""),
                    created_at,
                    cam["name"],
                    cam["resolution"],
                    cam["fps"],
                    cam["codec"],
                    cam["record_hour"],
                    cam["retention_days"],
                    cam["qty"],
                    doc.get("bandwidth", "N/A"),
                    doc.get("storage_tb", "N/A")
                ])

        filename_parts = ["redx_export"]
        if customer_name:
            filename_parts.append(f"cust_{customer_name[:20]}")
        if project_name:
            filename_parts.append(f"proj_{project_name[:20]}")
        if query:
            filename_parts.append(f"q_{query[:10]}")
        filename = "_".join(filename_parts) + ".xlsx"
        path = os.path.join(EXPORT_DIR, filename)    
        os.makedirs(EXPORT_DIR, exist_ok=True)
        wb.save(path)
        
        return FileResponse(path, filename=filename, headers={"Content-Disposition": f"attachment; filename={filename}"})
    
    except ValueError as e:
        raise HTTPException(status_code=422, detail=f"Invalid date format. Use YYYY-MM-DD: {str(e)}")
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail=f"Export Failed: {str(e)}"
        )



@app.get('/admin', response_class=HTMLResponse)
async def admin_dashboard(request: Request, current_user: UserInDB = Depends(get_superadmin)):
    try:        
        context = {
            "request": request,
            "local_ip": local_ip,
            "current_user": current_user
        }
        return templates.TemplateResponse("admin/users.html", context)
        
    except HTTPException as e:
        if e.status_code == status.HTTP_401_UNAUTHORIZED:
            context = {
                "request": request,
                "error": "Please login to access this page",
                "local_ip": local_ip
            }
            return templates.TemplateResponse("index1.html", context, status_code=401)
        context = {
            "request": request,
            "error": str(e.detail),
            "local_ip": local_ip
        }
        return templates.TemplateResponse("index1.html", context, status_code=e.status_code)
    except Exception as e:
        context = {
            "request": request,
            "error": "An error occurred while accessing this page",
            "local_ip": local_ip
        }
        return templates.TemplateResponse("index1.html", context, status_code=500)


@app.get("/users", response_model=List[UserOut])
async def get_all_users(current_user: UserInDB = Depends(get_superadmin)):
    if current_user.role == "admin":
        users_created_by_admin = list(users_collection.find({"created_by": current_user.username}))
        return users_created_by_admin
    users_created_by_superadmin = list(users_collection.find({"role": {"$ne": "superadmin"}}, {"_id": 0,"password": 0}))
    return users_created_by_superadmin

@app.post("/users", response_model=UserOut)
async def create_user(user: UserCreate, current_user: UserInDB = Depends(get_superadmin)):
    if users_collection.find_one({"username": user.username}):
        raise HTTPException(status_code=400, detail="Username already exists")
    
    hashed_pwd = get_password_hash(user.password)
    user_dict = user.dict()
    user_dict['hashed_password'] = hashed_pwd
    user_dict['is_active'] = user.is_active
    user_dict['created_by'] = current_user.username
    del user_dict['password']
    users_collection.insert_one(user_dict)
    return user_dict

@app.put("/users/{username}")
async def update_user(username: str, user: UserUpdate, current_user: UserInDB = Depends(get_superadmin)):
    update_data = {k: v for k, v in user.dict().items() if v is not None}
    if "password" in update_data:
        update_data["hashed_password"] = get_password_hash(update_data.pop("password"))
    
    result = users_collection.update_one({"username": username}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"msg": "User updated successfully"}

@app.delete("/users/{username}")
async def delete_user(username: str, current_user: UserInDB = Depends(get_superadmin)):
    result = users_collection.delete_one({"username": username})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"msg": "User deleted"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
